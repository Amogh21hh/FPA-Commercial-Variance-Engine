"""
Nova_FPA_Master_Model.xlsx builder — FAST-standard FP&A engine.
Tabs:
  1. Cover
  2. Assumptions       (macro + scenario toggle, yellow highlights)
  3. Mapping_Table     (account → category, fixes miscoded OPEX)
  4. Budget            (FY25 monthly budget by line)
  5. Actuals_Clean     (aggregated from raw ledger via SUMIFS)
  6. BvA_Dashboard     (variance + Price/Volume bridge + waterfall)
  7. Rolling_Forecast  (scenario-driven 12-month forecast)
"""
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT  = os.path.join(ROOT, "Nova_FPA_Master_Model.xlsx")

# ------------------------------------------------------------------
# Styles (per xlsx skill standards)
# ------------------------------------------------------------------
FONT_NAME = "Arial"
BLUE  = Font(name=FONT_NAME, color="0000FF")          # hardcoded inputs
BLACK = Font(name=FONT_NAME, color="000000")          # formulas
GREEN = Font(name=FONT_NAME, color="008000")          # cross-sheet links
RED   = Font(name=FONT_NAME, color="C00000")          # negative / external
HEADER = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11)
TITLE  = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=14)
BOLD   = Font(name=FONT_NAME, bold=True)
ITALIC = Font(name=FONT_NAME, italic=True, color="555555")

NAVY    = PatternFill("solid", start_color="1F2A44")
SLATE   = PatternFill("solid", start_color="2E3A59")
GREY    = PatternFill("solid", start_color="E9ECEF")
LIGHT   = PatternFill("solid", start_color="F5F7FA")
YELLOW  = PatternFill("solid", start_color="FFF59D")
GREEN_F = PatternFill("solid", start_color="C8E6C9")
RED_F   = PatternFill("solid", start_color="FFCDD2")

THIN = Side(style="thin", color="B0B7C3")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

CCY = '£#,##0;[Red](£#,##0);-'
PCT = '0.0%;[Red](0.0%);-'
NUM = '#,##0;[Red](#,##0);-'

# ------------------------------------------------------------------
# Load source CSVs
# ------------------------------------------------------------------
def load_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))

budget   = load_csv(os.path.join(ROOT, "FY25_Approved_Budget.csv"))
actuals  = load_csv(os.path.join(ROOT, "YTD_Actuals_Ledger.csv"))

# Unique budget lines (account, desc, cat)
seen = set()
lines = []
for r in budget:
    key = (r["Account_Code"], r["Description"], r["Category"])
    if key not in seen:
        seen.add(key); lines.append(key)
# Sort revenue → COGS → OPEX → Non-Cash → Finance
CAT_ORDER = {"Revenue":1,"COGS":2,"OPEX_Wages":3,"OPEX_Variable":4,"OPEX_Fixed":5,"Non-Cash":6,"Finance":7}
lines.sort(key=lambda x: (CAT_ORDER.get(x[2],9), x[0]))

PERIODS = [f"FY25-{m:02d}" for m in range(1,13)]
MONTH_LABELS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

wb = Workbook()

# ==================================================================
# TAB 1 — COVER
# ==================================================================
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for col in "BCDEFGHIJK":
    ws.column_dimensions[col].width = 14

ws.merge_cells("B2:K3")
ws["B2"] = "NOVA LEISURE GROUP"
ws["B2"].font = Font(name=FONT_NAME, size=24, bold=True, color="FFFFFF")
ws["B2"].alignment = CENTER
for r in (2,3):
    for c in range(2,12):
        ws.cell(row=r,column=c).fill = NAVY

ws.merge_cells("B4:K5")
ws["B4"] = "FP&A Rolling Forecast & Variance Engine  |  FY25 Close + FY26 Outlook"
ws["B4"].font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
ws["B4"].alignment = CENTER
for r in (4,5):
    for c in range(2,12):
        ws.cell(row=r,column=c).fill = SLATE

ws["B8"]  = "Prepared by:";  ws["B8"].font  = BOLD
ws["C8"]  = "Amogh H.H. — Commercial Finance"
ws["B9"]  = "Last refreshed:"; ws["B9"].font = BOLD
ws["C9"]  = "26 May 2026"
ws["B10"] = "Reporting standard:"; ws["B10"].font = BOLD
ws["C10"] = "FAST (Flexible · Appropriate · Structured · Transparent)"

ws["B13"] = "WORKBOOK MAP"; ws["B13"].font = TITLE; ws["B13"].fill = NAVY
ws.merge_cells("B13:K13"); ws["B13"].alignment = LEFT

tabs = [
    ("Assumptions",      "Macro + scenario toggle (Base / Bull / Bear). All yellow cells are user inputs."),
    ("Mapping_Table",    "Account-code dictionary. Catches miscoded OPEX before it hits the dashboard."),
    ("Budget",           "FY25 approved budget by line item × month. Source of truth for variances."),
    ("Actuals_Clean",    "Aggregated YTD actuals built from the raw ledger via SUMIFS + XLOOKUP."),
    ("BvA_Dashboard",    "Budget vs Actuals, EBITDA waterfall, and Price-vs-Volume revenue bridge."),
    ("Rolling_Forecast", "12-month rolling forecast. Sensitised to CPI + scenario toggle."),
]
for i,(t,d) in enumerate(tabs):
    r = 15+i
    ws.cell(row=r, column=2, value=t).font = BOLD
    ws.cell(row=r, column=2).fill = GREY
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=11)
    ws.cell(row=r, column=3, value=d).alignment = LEFT

ws["B23"] = "COLOUR CODING"; ws["B23"].font = TITLE; ws["B23"].fill = NAVY
ws.merge_cells("B23:K23")

key_rows = [
    ("Blue text",   "Hardcoded inputs the user can change",   BLUE,  None),
    ("Black text",  "Formulas / calculations",                 BLACK, None),
    ("Green text",  "Cross-sheet link",                        GREEN, None),
    ("Yellow fill", "Key assumption — review on refresh",      BLACK, YELLOW),
]
for i,(lbl,desc,fn,fill) in enumerate(key_rows):
    r = 25+i
    c = ws.cell(row=r, column=2, value=lbl); c.font = fn
    if fill: c.fill = fill
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=11)
    ws.cell(row=r, column=3, value=desc)

# ==================================================================
# TAB 2 — ASSUMPTIONS
# ==================================================================
asm = wb.create_sheet("Assumptions")
asm.sheet_view.showGridLines = False
asm.column_dimensions['A'].width = 2
asm.column_dimensions['B'].width = 32
for col in "CDEF":
    asm.column_dimensions[col].width = 14

asm["B2"] = "MACRO & SCENARIO ASSUMPTIONS"
asm["B2"].font = TITLE; asm["B2"].fill = NAVY
asm.merge_cells("B2:F2"); asm["B2"].alignment = LEFT

# Scenario selector
asm["B4"] = "Active Scenario:"; asm["B4"].font = BOLD
asm["C4"] = "Base"; asm["C4"].font = BLUE; asm["C4"].fill = YELLOW
asm["C4"].alignment = CENTER
dv = DataValidation(type="list", formula1='"Bear,Base,Bull"', allow_blank=False)
dv.add("C4")
asm.add_data_validation(dv)
asm["D4"] = '← Toggle scenario (drives Forecast tab)'
asm["D4"].font = ITALIC

# Scenario table
asm["B6"] = "Driver"; asm["C6"] = "Bear"; asm["D6"] = "Base"; asm["E6"] = "Bull"; asm["F6"] = "ACTIVE"
for c in "BCDEF":
    asm[f"{c}6"].font = HEADER; asm[f"{c}6"].fill = SLATE
    asm[f"{c}6"].alignment = CENTER

drivers = [
    ("CPI inflation (YoY)",     0.035, 0.027, 0.022, PCT),
    ("Wage bill uplift (YoY)",  0.070, 0.054, 0.040, PCT),
    ("Footfall growth (YoY)",  -0.060,-0.020, 0.015, PCT),
    ("Like-for-like sales",    -0.045,-0.010, 0.030, PCT),
    ("Energy unit cost change", 0.050,-0.030,-0.060, PCT),
    ("Avg spend / head growth", 0.010, 0.025, 0.040, PCT),
]
for i,(name,bear,base,bull,fmt) in enumerate(drivers):
    r = 7+i
    asm.cell(row=r,column=2,value=name).font = BLACK
    asm.cell(row=r,column=3,value=bear).font = BLUE; asm.cell(row=r,column=3).fill = YELLOW
    asm.cell(row=r,column=4,value=base).font = BLUE; asm.cell(row=r,column=4).fill = YELLOW
    asm.cell(row=r,column=5,value=bull).font = BLUE; asm.cell(row=r,column=5).fill = YELLOW
    asm.cell(row=r,column=6,value=f'=INDEX(C{r}:E{r},MATCH($C$4,$C$6:$E$6,0))').font = BLACK
    for c in "CDEF":
        asm[f"{c}{r}"].number_format = fmt
        asm[f"{c}{r}"].alignment = RIGHT

# Probability weights
asm["B14"] = "Probability weight"; asm["B14"].font = ITALIC
asm["C14"] = 0.25; asm["D14"] = 0.50; asm["E14"] = 0.25
for c in "CDE":
    asm[f"{c}14"].font = BLUE; asm[f"{c}14"].fill = YELLOW
    asm[f"{c}14"].number_format = PCT
asm["F14"] = '=SUM(C14:E14)'; asm["F14"].number_format = PCT
asm["F14"].font = BOLD

# Macro reference panel
asm["B17"] = "MACRO REFERENCE (Apr-26 prints, ONS / BRC)"
asm["B17"].font = HEADER; asm["B17"].fill = SLATE
asm.merge_cells("B17:F17")
macro = [
    ("Headline CPI YoY", "2.8%"),
    ("Core CPI YoY", "2.5%"),
    ("UK retail footfall YoY (Apr)", "-10.7%"),
    ("National Living Wage (age 21+)", "£12.21/hr"),
    ("BoE terminal Bank Rate (consensus)", "3.75% Q4-26"),
]
for i,(k,v) in enumerate(macro):
    r = 18+i
    asm.cell(row=r,column=2,value=k)
    asm.cell(row=r,column=3,value=v).font = BLUE
    asm.cell(row=r,column=3).alignment = LEFT
asm["B24"] = "Source: ONS CPI Bulletin Apr-26; BRC-Sensormatic Footfall Monitor; BoE MPR May-26"
asm["B24"].font = ITALIC

# ==================================================================
# TAB 3 — MAPPING TABLE
# ==================================================================
mp = wb.create_sheet("Mapping_Table")
mp.sheet_view.showGridLines = False
mp.column_dimensions['A'].width = 2
mp.column_dimensions['B'].width = 14
mp.column_dimensions['C'].width = 35
mp.column_dimensions['D'].width = 18
mp.column_dimensions['E'].width = 18

mp["B2"] = "CHART OF ACCOUNTS — MAPPING TABLE"
mp["B2"].font = TITLE; mp["B2"].fill = NAVY
mp.merge_cells("B2:E2")
mp["B4"] = "Used by Actuals_Clean (XLOOKUP) to map raw ledger codes to budget categories. Unknown codes default to 'UNMAPPED — REVIEW'."
mp["B4"].font = ITALIC; mp.merge_cells("B4:E4")

headers = ["Account_Code","Description","Category","EBITDA_Bucket"]
for i,h in enumerate(headers):
    c = mp.cell(row=6, column=2+i, value=h)
    c.font = HEADER; c.fill = SLATE; c.alignment = CENTER

def ebitda_bucket(cat):
    if cat == "Revenue": return "Revenue"
    if cat == "COGS": return "Gross Profit"
    if cat.startswith("OPEX"): return "EBITDA"
    return "Below EBITDA"

for i,(code,desc,cat) in enumerate(lines):
    r = 7+i
    mp.cell(row=r,column=2,value=code).font = BLUE
    mp.cell(row=r,column=3,value=desc)
    mp.cell(row=r,column=4,value=cat)
    mp.cell(row=r,column=5,value=ebitda_bucket(cat))

# Add miscoded "unmapped" handlers shown for transparency
extra_start = 7 + len(lines)
mp.cell(row=extra_start,column=2,value="6999").font = RED
mp.cell(row=extra_start,column=3,value="*** UNMAPPED OPEX — review ***").font = RED
mp.cell(row=extra_start,column=4,value="OPEX_Variable")
mp.cell(row=extra_start,column=5,value="EBITDA")
mp.cell(row=extra_start+1,column=2,value="9000").font = RED
mp.cell(row=extra_start+1,column=3,value="*** UNMAPPED SUSPENSE ***").font = RED
mp.cell(row=extra_start+1,column=4,value="OPEX_Variable")
mp.cell(row=extra_start+1,column=5,value="EBITDA")
mp.cell(row=extra_start+2,column=2,value="5099").font = RED
mp.cell(row=extra_start+2,column=3,value="*** UNMAPPED COGS ***").font = RED
mp.cell(row=extra_start+2,column=4,value="COGS")
mp.cell(row=extra_start+2,column=5,value="Gross Profit")

map_last = extra_start + 2

# Named range for mapping
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names["MapTable"] = DefinedName("MapTable", attr_text=f"Mapping_Table!$B$7:$E${map_last}")

# ==================================================================
# TAB 4 — BUDGET
# ==================================================================
bg = wb.create_sheet("Budget")
bg.sheet_view.showGridLines = False
bg.column_dimensions['A'].width = 2
bg.column_dimensions['B'].width = 10
bg.column_dimensions['C'].width = 32
bg.column_dimensions['D'].width = 16
for i,p in enumerate(MONTH_LABELS):
    bg.column_dimensions[get_column_letter(5+i)].width = 12
bg.column_dimensions[get_column_letter(5+12)].width = 14  # FY total

bg["B2"] = "FY25 APPROVED BUDGET (Monthly, £)"
bg["B2"].font = TITLE; bg["B2"].fill = NAVY
bg.merge_cells("B2:R2")

headers = ["Account","Description","Category"] + MONTH_LABELS + ["FY25 Total"]
for i,h in enumerate(headers):
    c = bg.cell(row=4, column=2+i, value=h)
    c.font = HEADER; c.fill = SLATE; c.alignment = CENTER

# Build budget by line and month
bdict = {}
for r in budget:
    bdict[(r["Account_Code"], r["Period"])] = float(r["Budget_GBP"])

for i,(code,desc,cat) in enumerate(lines):
    r = 5+i
    bg.cell(row=r,column=2,value=code).font = BLUE
    bg.cell(row=r,column=3,value=desc)
    bg.cell(row=r,column=4,value=cat)
    for j,p in enumerate(PERIODS):
        v = bdict.get((code,p), 0)
        cell = bg.cell(row=r, column=5+j, value=v)
        cell.font = BLUE
        cell.number_format = CCY
    fy_col = get_column_letter(5+12)
    bg.cell(row=r, column=5+12, value=f"=SUM(E{r}:P{r})").font = BLACK
    bg.cell(row=r, column=5+12).number_format = CCY

# Subtotals
sub_start = 5 + len(lines) + 1
def add_subtotal(r, label, cats):
    bg.cell(row=r, column=3, value=label).font = BOLD
    bg.cell(row=r, column=3).fill = GREY
    for j in range(13):
        col = get_column_letter(5+j)
        # SUMIFS over the lines' Category column (col D)
        formula = f'=SUMIFS({col}5:{col}{sub_start-2},$D$5:$D${sub_start-2},{_cat_string(cats)})'
        c = bg.cell(row=r, column=5+j, value=formula)
        c.font = BOLD; c.number_format = CCY; c.fill = GREY

def _cat_string(cats):
    # SUMIFS with multiple cats needs SUMPRODUCT; we cheat by joining with multiple SUMIFS
    return f'"{cats[0]}"' if len(cats) == 1 else None

# Use SUMPRODUCT for multi-category subtotals
def subtotal_formula(col, cats, last_row):
    parts = "+".join([f'SUMIFS({col}5:{col}{last_row},$D$5:$D${last_row},"{c}")' for c in cats])
    return f'={parts}'

line_last = 5 + len(lines) - 1
subs = [
    ("Total Revenue",      ["Revenue"]),
    ("Total COGS",         ["COGS"]),
    ("Gross Profit",       None),  # special
    ("Total OPEX",         ["OPEX_Wages","OPEX_Variable","OPEX_Fixed"]),
    ("EBITDA",             None),
    ("Depreciation",       ["Non-Cash"]),
    ("Interest",           ["Finance"]),
    ("Net Profit",         None),
]

sub_row = line_last + 2
sub_rows = {}
for label, cats in subs:
    sub_rows[label] = sub_row
    bg.cell(row=sub_row, column=3, value=label).font = BOLD
    bg.cell(row=sub_row, column=3).fill = GREY
    for j in range(13):
        col = get_column_letter(5+j)
        if cats:
            f = subtotal_formula(col, cats, line_last)
        elif label == "Gross Profit":
            f = f'={col}{sub_rows["Total Revenue"]}+{col}{sub_rows["Total COGS"]}'
        elif label == "EBITDA":
            f = f'={col}{sub_rows["Gross Profit"]}+{col}{sub_rows["Total OPEX"]}'
        elif label == "Net Profit":
            f = f'={col}{sub_rows["EBITDA"]}+{col}{sub_rows["Depreciation"]}+{col}{sub_rows["Interest"]}'
        c = bg.cell(row=sub_row, column=5+j, value=f)
        c.font = BOLD; c.number_format = CCY; c.fill = GREY
    sub_row += 1

bg.freeze_panes = "E5"

# ==================================================================
# TAB 5 — ACTUALS CLEAN
# ==================================================================
# Build an aggregated table from raw ledger via Python (so we can also
# show the SUMIFS formulas in the workbook for auditability). The raw
# ledger is dropped into a hidden "Ledger_Raw" tab so SUMIFS can run.
lr = wb.create_sheet("Ledger_Raw")
lr.sheet_state = "visible"  # keep visible for audit trail
lr.sheet_view.showGridLines = False
lr.column_dimensions['A'].width = 10
for c in "BCDEFGH":
    lr.column_dimensions[c].width = 16

ledger_headers = ["Txn_ID","Txn_Date","Period","Account_Code","Description","Site_Code","Department","Amount_GBP"]
for i,h in enumerate(ledger_headers):
    c = lr.cell(row=1, column=1+i, value=h)
    c.font = HEADER; c.fill = SLATE; c.alignment = CENTER

# Filter out fully blank rows for SUMIFS efficiency, but keep duplicates and miscodes
clean_rows = [r for r in actuals if r.get("Txn_ID","")]
for i,r in enumerate(clean_rows):
    row = i+2
    lr.cell(row=row, column=1, value=int(r["Txn_ID"]) if r["Txn_ID"] else None)
    lr.cell(row=row, column=2, value=r["Txn_Date"])
    lr.cell(row=row, column=3, value=r["Period"])
    lr.cell(row=row, column=4, value=r["Account_Code"])
    lr.cell(row=row, column=5, value=r["Description"])
    lr.cell(row=row, column=6, value=r["Site_Code"])
    lr.cell(row=row, column=7, value=r["Department"])
    try:
        lr.cell(row=row, column=8, value=float(r["Amount_GBP"]))
        lr.cell(row=row, column=8).number_format = CCY
    except (ValueError, TypeError):
        pass
ledger_last = len(clean_rows) + 1
lr.freeze_panes = "A2"

# Explicit ledger ranges (avoid named-range portability issues with recalc engine)
LEDGER_CODE   = f"Ledger_Raw!$D$2:$D${ledger_last}"
LEDGER_PERIOD = f"Ledger_Raw!$C$2:$C${ledger_last}"
LEDGER_AMOUNT = f"Ledger_Raw!$H$2:$H${ledger_last}"

ac = wb.create_sheet("Actuals_Clean")
ac.sheet_view.showGridLines = False
ac.column_dimensions['A'].width = 2
ac.column_dimensions['B'].width = 10
ac.column_dimensions['C'].width = 32
ac.column_dimensions['D'].width = 16
for i in range(13):
    ac.column_dimensions[get_column_letter(5+i)].width = 12
ac.column_dimensions[get_column_letter(5+12)].width = 14

ac["B2"] = "YTD ACTUALS — CLEANED & MAPPED (£)"
ac["B2"].font = TITLE; ac["B2"].fill = NAVY
ac.merge_cells("B2:R2")
ac["B3"] = "Built from Ledger_Raw via SUMIFS(LedgerAmount, LedgerCode, [code], LedgerPeriod, [period]). Category resolved via XLOOKUP to Mapping_Table."
ac["B3"].font = ITALIC
ac.merge_cells("B3:R3")

headers = ["Account","Description","Category"] + MONTH_LABELS + ["YTD Total"]
for i,h in enumerate(headers):
    c = ac.cell(row=5, column=2+i, value=h)
    c.font = HEADER; c.fill = SLATE; c.alignment = CENTER

for i,(code,desc,cat) in enumerate(lines):
    r = 6+i
    ac.cell(row=r,column=2,value=code).font = BLUE
    # INDEX/MATCH (portable across all Excel/LibreOffice versions)
    ac.cell(row=r,column=3,value=f'=IFERROR(INDEX(Mapping_Table!$C$7:$C${map_last},MATCH(B{r},Mapping_Table!$B$7:$B${map_last},0)),"UNMAPPED")').font = GREEN
    ac.cell(row=r,column=4,value=f'=IFERROR(INDEX(Mapping_Table!$D$7:$D${map_last},MATCH(B{r},Mapping_Table!$B$7:$B${map_last},0)),"UNMAPPED")').font = GREEN
    for j,p in enumerate(PERIODS):
        col = get_column_letter(5+j)
        f = f'=SUMIFS({LEDGER_AMOUNT},{LEDGER_CODE},B{r},{LEDGER_PERIOD},"{p}")'
        c = ac.cell(row=r, column=5+j, value=f)
        c.font = BLACK; c.number_format = CCY
    ac.cell(row=r, column=5+12, value=f"=SUM(E{r}:P{r})").font = BOLD
    ac.cell(row=r, column=5+12).number_format = CCY

# Unmapped catcher row
unmap_row = 6 + len(lines) + 1
ac.cell(row=unmap_row, column=3, value="UNMAPPED — Review (miscoded ledger entries)").font = Font(name=FONT_NAME, bold=True, color="C00000")
for j,p in enumerate(PERIODS):
    col = get_column_letter(5+j)
    f = f'=SUMIFS({LEDGER_AMOUNT},{LEDGER_PERIOD},"{p}")-SUM({col}6:{col}{6+len(lines)-1})'
    c = ac.cell(row=unmap_row, column=5+j, value=f)
    c.font = Font(name=FONT_NAME, color="C00000", bold=True); c.number_format = CCY
ac.cell(row=unmap_row, column=5+12, value=f"=SUM(E{unmap_row}:P{unmap_row})").font = Font(name=FONT_NAME, bold=True, color="C00000")
ac.cell(row=unmap_row, column=5+12).number_format = CCY

# Subtotals (mirror Budget)
sub_start_a = unmap_row + 2
line_last_a = 6 + len(lines) - 1
sub_rows_a = {}
sub_row = sub_start_a
for label, cats in subs:
    sub_rows_a[label] = sub_row
    ac.cell(row=sub_row, column=3, value=label).font = BOLD
    ac.cell(row=sub_row, column=3).fill = GREY
    for j in range(13):
        col = get_column_letter(5+j)
        if cats:
            f = "=" + "+".join([f'SUMIFS({col}6:{col}{line_last_a},$D$6:$D${line_last_a},"{c}")' for c in cats])
        elif label == "Gross Profit":
            f = f'={col}{sub_rows_a["Total Revenue"]}+{col}{sub_rows_a["Total COGS"]}'
        elif label == "EBITDA":
            f = f'={col}{sub_rows_a["Gross Profit"]}+{col}{sub_rows_a["Total OPEX"]}'
        elif label == "Net Profit":
            f = f'={col}{sub_rows_a["EBITDA"]}+{col}{sub_rows_a["Depreciation"]}+{col}{sub_rows_a["Interest"]}'
        c = ac.cell(row=sub_row, column=5+j, value=f)
        c.font = BOLD; c.number_format = CCY; c.fill = GREY
    sub_row += 1

ac.freeze_panes = "E6"

# ==================================================================
# TAB 6 — BvA DASHBOARD
# ==================================================================
bv = wb.create_sheet("BvA_Dashboard")
bv.sheet_view.showGridLines = False
bv.column_dimensions['A'].width = 2
bv.column_dimensions['B'].width = 32
for c in "CDEFGHI":
    bv.column_dimensions[c].width = 16

bv["B2"] = "BUDGET vs ACTUALS — FY25 YTD"
bv["B2"].font = TITLE; bv["B2"].fill = NAVY
bv.merge_cells("B2:I2")

bv["B4"] = "P&L SUMMARY (£)"
bv["B4"].font = HEADER; bv["B4"].fill = SLATE
bv.merge_cells("B4:I4")

bv["B5"] = "Metric"; bv["C5"] = "Budget"; bv["D5"] = "Actual"
bv["E5"] = "Variance £"; bv["F5"] = "Variance %"; bv["G5"] = "Status"
for c in "BCDEFG":
    bv[f"{c}5"].font = HEADER; bv[f"{c}5"].fill = SLATE; bv[f"{c}5"].alignment = CENTER

metric_rows = [
    ("Total Revenue", sub_rows["Total Revenue"], sub_rows_a["Total Revenue"], False),
    ("Total COGS",    sub_rows["Total COGS"],    sub_rows_a["Total COGS"], True),
    ("Gross Profit",  sub_rows["Gross Profit"],  sub_rows_a["Gross Profit"], False),
    ("Total OPEX",    sub_rows["Total OPEX"],    sub_rows_a["Total OPEX"], True),
    ("EBITDA",        sub_rows["EBITDA"],        sub_rows_a["EBITDA"], False),
    ("Net Profit",    sub_rows["Net Profit"],    sub_rows_a["Net Profit"], False),
]

fy_col_b = get_column_letter(5+12)  # Q (Budget FY total)
fy_col_a = get_column_letter(5+12)  # Q (Actuals FY total)

for i,(name,b_row,a_row,is_cost) in enumerate(metric_rows):
    r = 6+i
    bv.cell(row=r, column=2, value=name).font = BOLD if name in ("EBITDA","Gross Profit","Net Profit") else BLACK
    bv.cell(row=r, column=3, value=f"=Budget!{fy_col_b}{b_row}").font = GREEN
    bv.cell(row=r, column=4, value=f"=Actuals_Clean!{fy_col_a}{a_row}").font = GREEN
    bv.cell(row=r, column=5, value=f"=D{r}-C{r}").font = BLACK
    bv.cell(row=r, column=6, value=f'=IFERROR((D{r}-C{r})/ABS(C{r}),0)').font = BLACK
    # Costs in this model are stored as NEGATIVES.
    # → Cost variance (Actual - Budget) > 0 means actual is LESS negative = under-spend = Favourable.
    # → Revenue/profit variance > 0 is straightforwardly Favourable.
    bv.cell(row=r, column=7, value=f'=IF(E{r}>=0,"Favourable","Adverse")')
    for c in "CDE":
        bv[f"{c}{r}"].number_format = CCY
    bv[f"F{r}"].number_format = PCT
    bv[f"G{r}"].alignment = CENTER
    if name in ("EBITDA","Gross Profit","Net Profit"):
        for c in "BCDEFG":
            bv[f"{c}{r}"].fill = LIGHT
            bv[f"{c}{r}"].font = BOLD if c == "B" else Font(name=FONT_NAME, bold=True)

# Conditional format Status
fav_rule = CellIsRule(operator='equal', formula=['"Favourable"'], fill=GREEN_F)
adv_rule = CellIsRule(operator='equal', formula=['"Adverse"'], fill=RED_F)
bv.conditional_formatting.add(f"G6:G{6+len(metric_rows)-1}", fav_rule)
bv.conditional_formatting.add(f"G6:G{6+len(metric_rows)-1}", adv_rule)

# ---------- EBITDA Waterfall data ----------
wf_start = 14
bv.cell(row=wf_start, column=2, value="EBITDA BRIDGE (Budget → Actual, £)").font = HEADER
bv.cell(row=wf_start, column=2).fill = SLATE
bv.merge_cells(start_row=wf_start, start_column=2, end_row=wf_start, end_column=9)

# Bridge categories
bv.cell(row=wf_start+1, column=2, value="Bucket").font = HEADER; bv.cell(row=wf_start+1, column=2).fill = SLATE
bv.cell(row=wf_start+1, column=3, value="Δ £").font = HEADER; bv.cell(row=wf_start+1, column=3).fill = SLATE
bv.cell(row=wf_start+1, column=4, value="Running £").font = HEADER; bv.cell(row=wf_start+1, column=4).fill = SLATE
for col in "BCD":
    bv[f"{col}{wf_start+1}"].alignment = CENTER

# Bridge rows
ebitda_b_row = sub_rows["EBITDA"]
ebitda_a_row = sub_rows_a["EBITDA"]
rev_b_row = sub_rows["Total Revenue"]; rev_a_row = sub_rows_a["Total Revenue"]
cogs_b_row = sub_rows["Total COGS"]; cogs_a_row = sub_rows_a["Total COGS"]
wages_a_subline = sub_rows_a["Total OPEX"]  # we'll break OPEX

# Compute OPEX wages / variable / fixed via SUMIFS on Actuals_Clean
bridge_rows = [
    ("Budget EBITDA",      f"=Budget!{fy_col_b}{ebitda_b_row}", "anchor"),
    ("Revenue Δ",          f"=Actuals_Clean!{fy_col_a}{rev_a_row}-Budget!{fy_col_b}{rev_b_row}", "delta"),
    ("COGS Δ",             f"=Actuals_Clean!{fy_col_a}{cogs_a_row}-Budget!{fy_col_b}{cogs_b_row}", "delta"),
    ("Wages Δ",            f'=SUMIFS(Actuals_Clean!{fy_col_a}6:{fy_col_a}{line_last_a},Actuals_Clean!$D$6:$D${line_last_a},"OPEX_Wages")'
                            f'-SUMIFS(Budget!{fy_col_b}5:{fy_col_b}{line_last},Budget!$D$5:$D${line_last},"OPEX_Wages")', "delta"),
    ("Variable OPEX Δ",    f'=SUMIFS(Actuals_Clean!{fy_col_a}6:{fy_col_a}{line_last_a},Actuals_Clean!$D$6:$D${line_last_a},"OPEX_Variable")'
                            f'-SUMIFS(Budget!{fy_col_b}5:{fy_col_b}{line_last},Budget!$D$5:$D${line_last},"OPEX_Variable")', "delta"),
    ("Fixed OPEX Δ",       f'=SUMIFS(Actuals_Clean!{fy_col_a}6:{fy_col_a}{line_last_a},Actuals_Clean!$D$6:$D${line_last_a},"OPEX_Fixed")'
                            f'-SUMIFS(Budget!{fy_col_b}5:{fy_col_b}{line_last},Budget!$D$5:$D${line_last},"OPEX_Fixed")', "delta"),
    ("Actual EBITDA",      None, "anchor_end"),
]
# NOTE: Unmapped/miscoded entries are shown as a MEMO line below the bridge
# (not added to the running total) because they are not yet in EBITDA — they
# sit in suspense awaiting reclassification by the FP&A team.

run = wf_start + 2
for i,(name, formula, kind) in enumerate(bridge_rows):
    r = run + i
    bv.cell(row=r, column=2, value=name).font = BOLD if kind.startswith("anchor") else BLACK
    if formula:
        bv.cell(row=r, column=3, value=formula).font = BLACK
    else:
        bv.cell(row=r, column=3, value=f"=Actuals_Clean!{fy_col_a}{ebitda_a_row}-D{r-1}").font = BLACK
    # Running total
    if kind == "anchor":
        bv.cell(row=r, column=4, value=f"=C{r}").font = BOLD
    elif kind == "anchor_end":
        bv.cell(row=r, column=4, value=f"=D{r-1}+C{r}").font = BOLD
    else:
        bv.cell(row=r, column=4, value=f"=D{r-1}+C{r}").font = BLACK
    for c in "CD":
        bv[f"{c}{r}"].number_format = CCY
    if kind.startswith("anchor"):
        for c in "BCD":
            bv[f"{c}{r}"].fill = LIGHT

# Memo row below the bridge (NOT in chart)
memo_r = run + len(bridge_rows) + 1
bv.cell(row=memo_r, column=2, value="Memo: Unmapped / miscoded (suspense)").font = Font(name=FONT_NAME, italic=True, color="C00000")
bv.cell(row=memo_r, column=3, value=f'=SUM(Actuals_Clean!E{unmap_row}:P{unmap_row})').font = Font(name=FONT_NAME, italic=True, color="C00000")
bv.cell(row=memo_r, column=3).number_format = CCY
bv.cell(row=memo_r, column=4, value="Awaits reclassification — see Actuals_Clean UNMAPPED row").font = ITALIC

# Bar chart for waterfall (approximate — stacked positive/negative)
chart_top = run
chart_bot = run + len(bridge_rows) - 1
chart = BarChart()
chart.type = "col"
chart.style = 11
chart.title = "EBITDA Bridge: Budget → Actual"
chart.y_axis.title = "£"
chart.x_axis.title = None
data = Reference(bv, min_col=3, min_row=chart_top-1, max_row=chart_bot, max_col=3)
cats = Reference(bv, min_col=2, min_row=chart_top, max_row=chart_bot)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 9
chart.width = 18
chart.dataLabels = DataLabelList(showVal=True)
bv.add_chart(chart, "F14")

# ---------- Price vs Volume bridge ----------
pv_start = chart_bot + 3
bv.cell(row=pv_start, column=2, value="REVENUE BRIDGE — PRICE vs VOLUME").font = HEADER
bv.cell(row=pv_start, column=2).fill = SLATE
bv.merge_cells(start_row=pv_start, start_column=2, end_row=pv_start, end_column=6)

# Pull footfall totals for FY25 only (Jan-Dec 2025)
# Compute YTD visitors and avg spend from footfall CSV (precompute)
foot = load_csv(os.path.join(ROOT, "Daily_Site_Footfall.csv"))
fy25 = [r for r in foot if r["Date"].startswith("2025")]
total_visitors_fy25 = sum(int(r["Visitors"]) for r in fy25)
avg_spend_fy25 = sum(float(r["Avg_Spend_per_Head_GBP"]) * int(r["Visitors"]) for r in fy25) / total_visitors_fy25

# Budget assumes: budget revenue / budget visitors at budget spend.
# For the bridge we'll use proxy: budget_visitors implied by budget_revenue / budget_avg_spend.
budget_avg_spend = 14.50  # assumed budget spend/head

bv.cell(row=pv_start+1, column=2, value="Metric").font = HEADER; bv.cell(row=pv_start+1, column=2).fill = SLATE
bv.cell(row=pv_start+1, column=3, value="Budget").font = HEADER; bv.cell(row=pv_start+1, column=3).fill = SLATE
bv.cell(row=pv_start+1, column=4, value="Actual").font = HEADER; bv.cell(row=pv_start+1, column=4).fill = SLATE
bv.cell(row=pv_start+1, column=5, value="Variance").font = HEADER; bv.cell(row=pv_start+1, column=5).fill = SLATE
bv.cell(row=pv_start+1, column=6, value="% Var").font = HEADER; bv.cell(row=pv_start+1, column=6).fill = SLATE
for c in "BCDEF":
    bv[f"{c}{pv_start+1}"].alignment = CENTER

# Total Revenue line
bv.cell(row=pv_start+2, column=2, value="Total Revenue (£)")
bv.cell(row=pv_start+2, column=3, value=f"=Budget!{fy_col_b}{rev_b_row}").font = GREEN
bv.cell(row=pv_start+2, column=4, value=f"=Actuals_Clean!{fy_col_a}{rev_a_row}").font = GREEN
bv.cell(row=pv_start+2, column=5, value=f"=D{pv_start+2}-C{pv_start+2}")
bv.cell(row=pv_start+2, column=6, value=f"=E{pv_start+2}/ABS(C{pv_start+2})")
for c in "CDE": bv[f"{c}{pv_start+2}"].number_format = CCY
bv[f"F{pv_start+2}"].number_format = PCT

# Volume: Total visitors
bv.cell(row=pv_start+3, column=2, value="Volume (visitors)")
bv.cell(row=pv_start+3, column=3, value=f"=C{pv_start+2}/D{pv_start+5}")  # implied
bv.cell(row=pv_start+3, column=4, value=total_visitors_fy25)
bv.cell(row=pv_start+3, column=4).font = BLUE
bv.cell(row=pv_start+3, column=5, value=f"=D{pv_start+3}-C{pv_start+3}")
bv.cell(row=pv_start+3, column=6, value=f"=E{pv_start+3}/C{pv_start+3}")
for c in "CDE": bv[f"{c}{pv_start+3}"].number_format = NUM
bv[f"F{pv_start+3}"].number_format = PCT

# Avg spend per head
bv.cell(row=pv_start+4, column=2, value="Price (£/head)")
bv.cell(row=pv_start+4, column=5, value=f"=D{pv_start+5}-C{pv_start+5}")
bv.cell(row=pv_start+4, column=6, value=f"=E{pv_start+4}/C{pv_start+5}")
bv[f"F{pv_start+4}"].number_format = PCT
bv.cell(row=pv_start+5, column=2, value="  Avg spend / head (£)")
bv.cell(row=pv_start+5, column=3, value=budget_avg_spend)
bv.cell(row=pv_start+5, column=3).font = BLUE; bv.cell(row=pv_start+5, column=3).fill = YELLOW
bv.cell(row=pv_start+5, column=4, value=round(avg_spend_fy25, 2))
bv.cell(row=pv_start+5, column=4).font = BLUE
for c in "CDE": bv[f"{c}{pv_start+5}"].number_format = '£#,##0.00'

# Variance decomposition (Price × Vol)
bv.cell(row=pv_start+7, column=2, value="VOLUME EFFECT (ΔVol × Budget Price)").font = BOLD
bv.cell(row=pv_start+7, column=3, value=f"=(D{pv_start+3}-C{pv_start+3})*C{pv_start+5}")
bv.cell(row=pv_start+7, column=3).font = BOLD; bv.cell(row=pv_start+7, column=3).number_format = CCY

bv.cell(row=pv_start+8, column=2, value="PRICE EFFECT (ΔPrice × Actual Vol)").font = BOLD
bv.cell(row=pv_start+8, column=3, value=f"=(D{pv_start+5}-C{pv_start+5})*D{pv_start+3}")
bv.cell(row=pv_start+8, column=3).font = BOLD; bv.cell(row=pv_start+8, column=3).number_format = CCY

bv.cell(row=pv_start+9, column=2, value="TOTAL DECOMPOSED").font = BOLD
bv.cell(row=pv_start+9, column=2).fill = GREY
bv.cell(row=pv_start+9, column=3, value=f"=C{pv_start+7}+C{pv_start+8}")
bv.cell(row=pv_start+9, column=3).font = BOLD; bv.cell(row=pv_start+9, column=3).number_format = CCY
bv.cell(row=pv_start+9, column=3).fill = GREY

# ==================================================================
# TAB 7 — ROLLING FORECAST
# ==================================================================
rf = wb.create_sheet("Rolling_Forecast")
rf.sheet_view.showGridLines = False
rf.column_dimensions['A'].width = 2
rf.column_dimensions['B'].width = 32
rf.column_dimensions['C'].width = 14
for i in range(12):
    rf.column_dimensions[get_column_letter(4+i)].width = 12
rf.column_dimensions[get_column_letter(4+12)].width = 14

rf["B2"] = "ROLLING 12-MONTH FORECAST (Jun-26 → May-27)"
rf["B2"].font = TITLE; rf["B2"].fill = NAVY
rf.merge_cells("B2:Q2")

rf["B4"] = "Active scenario:"; rf["B4"].font = BOLD
rf["C4"] = "=Assumptions!C4"; rf["C4"].font = GREEN; rf["C4"].fill = YELLOW; rf["C4"].alignment = CENTER

# Pull active drivers
rf["E4"] = "CPI:";       rf["F4"] = "=Assumptions!F7";  rf["F4"].number_format = PCT; rf["F4"].font = GREEN
rf["G4"] = "Wages:";     rf["H4"] = "=Assumptions!F8";  rf["H4"].number_format = PCT; rf["H4"].font = GREEN
rf["I4"] = "Footfall:";  rf["J4"] = "=Assumptions!F9";  rf["J4"].number_format = PCT; rf["J4"].font = GREEN
rf["K4"] = "LfL:";       rf["L4"] = "=Assumptions!F10"; rf["L4"].number_format = PCT; rf["L4"].font = GREEN
rf["M4"] = "Energy:";    rf["N4"] = "=Assumptions!F11"; rf["N4"].number_format = PCT; rf["N4"].font = GREEN
for col in "EGIKM": rf[f"{col}4"].font = BOLD

# Header row
fc_months = ["Jun-26","Jul-26","Aug-26","Sep-26","Oct-26","Nov-26","Dec-26","Jan-27","Feb-27","Mar-27","Apr-27","May-27"]
rf.cell(row=6, column=2, value="Line").font = HEADER; rf.cell(row=6, column=2).fill = SLATE
rf.cell(row=6, column=3, value="FY25 Actual").font = HEADER; rf.cell(row=6, column=3).fill = SLATE
for i,m in enumerate(fc_months):
    c = rf.cell(row=6, column=4+i, value=m)
    c.font = HEADER; c.fill = SLATE; c.alignment = CENTER
rf.cell(row=6, column=4+12, value="NTM Total").font = HEADER
rf.cell(row=6, column=4+12).fill = SLATE
rf.cell(row=6, column=4+12).alignment = CENTER

# Forecast lines — driver mapping
# Revenue: grow by LfL + footfall effect, monthly seasonality from FY25 actuals
# COGS: scale with Revenue (margin constant from FY25)
# OPEX_Wages: +Wages%
# OPEX_Variable: +CPI (except energy which uses Energy%)
# OPEX_Fixed: +CPI
# Non-Cash / Finance: flat

fc_lines = [
    ("Revenue",        sub_rows_a["Total Revenue"], "revenue"),
    ("COGS",           sub_rows_a["Total COGS"],    "cogs"),
    ("Gross Profit",   None, "gp"),
    ("Wages",          None, "wages"),
    ("Variable OPEX",  None, "varopex"),
    ("Fixed OPEX",     None, "fixopex"),
    ("Total OPEX",     None, "opex"),
    ("EBITDA",         None, "ebitda"),
]

# Pre-compute the FY25 base values (these will use formulas referencing Actuals)
fc_row_map = {}
for i,(name, a_row, kind) in enumerate(fc_lines):
    r = 7+i
    fc_row_map[kind] = r
    rf.cell(row=r, column=2, value=name).font = BOLD if name in ("EBITDA","Gross Profit","Total OPEX") else BLACK

# FY25 base
rev_act_cell  = f"Actuals_Clean!{fy_col_a}{sub_rows_a['Total Revenue']}"
cogs_act_cell = f"Actuals_Clean!{fy_col_a}{sub_rows_a['Total COGS']}"

# Wages / Var OPEX / Fix OPEX from Actuals via SUMIFS
wages_act  = f'=SUMIFS(Actuals_Clean!{fy_col_a}6:{fy_col_a}{line_last_a},Actuals_Clean!$D$6:$D${line_last_a},"OPEX_Wages")'
varop_act  = f'=SUMIFS(Actuals_Clean!{fy_col_a}6:{fy_col_a}{line_last_a},Actuals_Clean!$D$6:$D${line_last_a},"OPEX_Variable")'
fixop_act  = f'=SUMIFS(Actuals_Clean!{fy_col_a}6:{fy_col_a}{line_last_a},Actuals_Clean!$D$6:$D${line_last_a},"OPEX_Fixed")'

rf.cell(row=fc_row_map["revenue"],  column=3, value=f"={rev_act_cell}").font = GREEN
rf.cell(row=fc_row_map["cogs"],     column=3, value=f"={cogs_act_cell}").font = GREEN
rf.cell(row=fc_row_map["gp"],       column=3, value=f"=C{fc_row_map['revenue']}+C{fc_row_map['cogs']}").font = BOLD
rf.cell(row=fc_row_map["wages"],    column=3, value=wages_act).font = GREEN
rf.cell(row=fc_row_map["varopex"],  column=3, value=varop_act).font = GREEN
rf.cell(row=fc_row_map["fixopex"],  column=3, value=fixop_act).font = GREEN
rf.cell(row=fc_row_map["opex"],     column=3, value=f"=C{fc_row_map['wages']}+C{fc_row_map['varopex']}+C{fc_row_map['fixopex']}").font = BOLD
rf.cell(row=fc_row_map["ebitda"],   column=3, value=f"=C{fc_row_map['gp']}+C{fc_row_map['opex']}").font = BOLD

# Seasonality factors for forecast (Jun-26 .. May-27 maps to months 6..12, 1..5)
seas = [1.05,1.20,1.25,1.00,0.95,1.05,1.40,0.85,0.82,0.90,1.00,1.10]
# Monthly base = FY25 annual /12 * seasonality
for i,month in enumerate(fc_months):
    col = get_column_letter(4+i)
    s = seas[i]
    # Revenue
    rf.cell(row=fc_row_map["revenue"], column=4+i,
            value=f"=C{fc_row_map['revenue']}/12*{s}*(1+$L$4)*(1+$J$4)").font = BLACK
    # COGS scales with revenue at same FY25 margin (IFERROR guards startup states)
    rf.cell(row=fc_row_map["cogs"], column=4+i,
            value=f"=IFERROR({col}{fc_row_map['revenue']}*(C{fc_row_map['cogs']}/C{fc_row_map['revenue']}),0)").font = BLACK
    # GP
    rf.cell(row=fc_row_map["gp"], column=4+i,
            value=f"={col}{fc_row_map['revenue']}+{col}{fc_row_map['cogs']}").font = BOLD
    # Wages (flat seasonality)
    rf.cell(row=fc_row_map["wages"], column=4+i,
            value=f"=C{fc_row_map['wages']}/12*(1+$H$4)").font = BLACK
    # Variable OPEX — CPI uplift, light seasonality
    rf.cell(row=fc_row_map["varopex"], column=4+i,
            value=f"=C{fc_row_map['varopex']}/12*(1+$F$4)*{s:.2f}").font = BLACK
    # Fixed OPEX — CPI uplift only
    rf.cell(row=fc_row_map["fixopex"], column=4+i,
            value=f"=C{fc_row_map['fixopex']}/12*(1+$F$4)").font = BLACK
    rf.cell(row=fc_row_map["opex"], column=4+i,
            value=f"={col}{fc_row_map['wages']}+{col}{fc_row_map['varopex']}+{col}{fc_row_map['fixopex']}").font = BOLD
    rf.cell(row=fc_row_map["ebitda"], column=4+i,
            value=f"={col}{fc_row_map['gp']}+{col}{fc_row_map['opex']}").font = BOLD

# NTM totals
ntm_col = get_column_letter(4+12)
for k,r in fc_row_map.items():
    rf.cell(row=r, column=4+12, value=f"=SUM(D{r}:O{r})").font = BOLD
# Format
for r in fc_row_map.values():
    for c in range(3, 4+13):
        rf.cell(row=r, column=c).number_format = CCY
    if rf.cell(row=r, column=2).value in ("EBITDA","Gross Profit","Total OPEX"):
        for c in range(2, 4+13):
            rf.cell(row=r, column=c).fill = LIGHT

# Forecast chart: EBITDA monthly
chart2 = LineChart()
chart2.title = "Forecast EBITDA — Next 12 Months"
chart2.y_axis.title = "EBITDA (£)"
chart2.x_axis.title = "Month"
data2 = Reference(rf, min_col=4, max_col=4+11, min_row=fc_row_map["ebitda"], max_row=fc_row_map["ebitda"])
cats2 = Reference(rf, min_col=4, max_col=4+11, min_row=6, max_row=6)
chart2.add_data(data2, titles_from_data=False)
chart2.set_categories(cats2)
chart2.height = 9
chart2.width = 18
rf.add_chart(chart2, "B18")

# Predictive overlay note
rf["B30"] = "PREDICTIVE OVERLAY (from R Time-Series Model)"
rf["B30"].font = HEADER; rf["B30"].fill = SLATE
rf.merge_cells("B30:Q30")
rf["B31"] = "Footfall forecast from Predictive_Demand_Model/forecast_output.csv (ARIMA / Holt-Winters)."
rf["B31"].font = ITALIC
rf["B32"] = "Quarterly visitors forecast (Jun-Aug 26)"
rf["B33"] = "MAPE:"; rf["B33"].font = BOLD
rf["C33"] = "see Predictive_Demand_Model/model_metrics.txt"
rf["C33"].font = ITALIC

rf.freeze_panes = "D7"

# ==================================================================
# Reorder tabs and save
# ==================================================================
order = ["Cover","Assumptions","Mapping_Table","Budget","Actuals_Clean","BvA_Dashboard","Rolling_Forecast","Ledger_Raw"]
wb._sheets = [wb[t] for t in order]

wb.save(OUT)
print(f"Wrote {OUT}")
 see Predictive_Demand_Model/model_metrics.txt"
rf["B33"].font = ITALIC

rf.freeze_panes = "D7"

# ==================================================================
# Reorder tabs and save
# ==================================================================
order = ["Cover","Assumptions","Mapping_Table","Budget","Actuals_Clean","BvA_Dashboard","Rolling_Forecast","Ledger_Raw"]
wb._sheets = [wb[t] for t in order]

wb.save(OUT)
print(f"Wrote {OUT}")


wb.save(OUT)
print(f"Wrote {OUT}")
